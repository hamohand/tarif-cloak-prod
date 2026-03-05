#!/usr/bin/env python3
"""
Convertit un fichier Excel CN (Combined Nomenclature) en CSV pour import PostgreSQL.
Extrait uniquement les codes à 8 chiffres → table position8_dz(code, description).

Usage :
    python cn_to_position8_csv.py <fichier.xlsx> [output.csv]

Exemple :
    python cn_to_position8_csv.py "CN 2026 official texts.xlsx"
    python cn_to_position8_csv.py "CN subheadings 2026.xlsx" position8_dz.csv

Dépendances :
    pip install openpyxl
"""

import sys
import csv
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installez openpyxl : pip install openpyxl")
    sys.exit(1)


def normalize_code(raw) -> str | None:
    """
    Retourne le code CN à 8 chiffres bruts (ex: '87031010').
    Accepte '8703 10 10', '87031010', '8703101000' (TARIC 10 chiffres → tronque les 2 derniers).
    Retourne None si ce n'est pas un code à 8 chiffres.
    """
    if raw is None:
        return None
    s = str(raw).strip().replace(" ", "").replace(".", "").replace("-", "")
    # Supprimer tout sauf les chiffres
    digits = re.sub(r"[^0-9]", "", s)
    if len(digits) == 10 and digits.endswith("00"):
        # Code TARIC 10 chiffres avec 00 final → niveau CN à 8 chiffres
        digits = digits[:8]
    if len(digits) == 8:
        return digits
    return None


def clean_description(raw) -> str:
    """Nettoie la description : supprime les espaces multiples et les tirets de début."""
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"\s+", " ", s)
    # Certains fichiers CN préfixent avec des tirets ou des points pour l'indentation
    s = s.lstrip("- ").strip()
    return s


def inspect_sheet(ws):
    """Affiche les 5 premières lignes pour aider l'utilisateur à identifier les colonnes."""
    print("\n--- Aperçu des 5 premières lignes ---")
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        print(f"  Ligne {i+1}: {row}")
    print("-------------------------------------\n")


def find_code_desc_columns(ws):
    """
    Tente de deviner automatiquement les colonnes code et description
    en cherchant un en-tête ou en analysant les données.
    Retourne (col_code_idx, col_desc_idx) basés sur 0, ou None si indétectable.
    """
    # Chercher dans la première ligne un en-tête ressemblant à code/description
    headers = [str(c).lower() if c else "" for c in next(ws.iter_rows(max_row=1, values_only=True))]
    code_idx = None
    desc_idx = None

    for i, h in enumerate(headers):
        if any(k in h for k in ["code", "cn", "nummer", "numero"]):
            code_idx = i
        if any(k in h for k in ["desc", "text", "designation", "libelle", "nom"]):
            desc_idx = i

    if code_idx is not None and desc_idx is not None:
        print(f"  En-têtes détectés : code=colonne {code_idx+1}, description=colonne {desc_idx+1}")
        return code_idx, desc_idx

    # Pas d'en-tête clair → essai colonne 0 = code, colonne 1 = description
    return 0, 1


def convert(xlsx_path: str, output_path: str):
    print(f"Ouverture de : {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Prendre le premier onglet
    ws = wb.active
    print(f"Onglet actif : '{ws.title}' — {ws.max_row} lignes × {ws.max_column} colonnes")

    inspect_sheet(ws)
    code_col, desc_col = find_code_desc_columns(ws)

    rows_out = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(code_col, desc_col):
            continue
        code = normalize_code(row[code_col])
        if code is None:
            skipped += 1
            continue
        desc = clean_description(row[desc_col])
        if not desc:
            skipped += 1
            continue
        rows_out.append((code, desc))

    wb.close()

    # Dédoublonnage par code (premier rencontré)
    seen = {}
    unique_rows = []
    for code, desc in rows_out:
        if code not in seen:
            seen[code] = True
            unique_rows.append((code, desc))

    print(f"\nRésultats :")
    print(f"  Codes à 8 chiffres extraits : {len(unique_rows)}")
    print(f"  Lignes ignorées (pas un code 8 chiffres) : {skipped}")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(["code", "description"])
        writer.writerows(unique_rows)

    print(f"\nFichier CSV créé : {output_path}")
    print(f"\nImport PostgreSQL (sur le VPS) :")
    print(f"  psql -U muhend -d hscode-app-db -c \"\\copy position8_dz(code,description) FROM '/tmp/{Path(output_path).name}' CSV HEADER;\"")
    print(f"\nOu avec upsert (si la table contient déjà des données) :")
    print(f"  psql -U muhend -d hscode-app-db < import_position8.sql")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    xlsx = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "position8_dz.csv"

    if not Path(xlsx).exists():
        print(f"Fichier introuvable : {xlsx}")
        sys.exit(1)

    convert(xlsx, out)
